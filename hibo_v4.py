import pandas as pd
import numpy as np
import os
from datetime import datetime
from sklearn.linear_model import LinearRegression
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- BULUT UYUMLU AYARLAR ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'DATAson')
OUTPUT_DIR = BASE_DIR

if not os.path.exists(DATA_DIR): os.makedirs(DATA_DIR)

# --- YARDIMCI FONKSİYONLAR ---
def load_stock_df(file_path):
    try:
        df = pd.read_excel(file_path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        col_map = {"DATE": ["DATE"], "CLOSING_TL": ["CLOSING_TL","CLOSE"]}
        for t, aliases in col_map.items():
            if t not in df.columns:
                for a in aliases:
                    if a in df.columns: df.rename(columns={a: t}, inplace=True); break
        return df if "CLOSING_TL" in df.columns else None
    except: return None

def calculate_ema(data, period): return data.ewm(span=period, adjust=False).mean()
def calculate_wma(data, period):
    weights = np.arange(1, period + 1)
    return data.rolling(period).apply(lambda x: np.dot(x, weights) / weights.sum(), raw=True)
def calculate_ma_type(data, p):
    e1=calculate_ema(data,p); e2=calculate_ema(e1,p); e3=calculate_ema(e2,p)
    return 3*e1 - 3*e2 + e3

def main():
    print("Hibrit V4 (Full) Taraması Başlıyor...")
    files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')]
    results = []
    
    # Kanal Periyotları
    check_periods = [55, 89, 144, 233, 370, 610]
    ema_periods = [8, 13, 21, 34, 55, 89, 144, 233]
    
    for f in files:
        try:
            df = load_stock_df(os.path.join(DATA_DIR, f))
            if df is None or len(df) < 610: continue
            
            close = df['CLOSING_TL']
            curr = close.iloc[-1]
            hisse = f.replace('.xlsx','')
            
            # 1. EMA KONTROLÜ
            emas = {p: calculate_ema(close, p).iloc[-1] for p in ema_periods}
            
            # IDEAL UP: Fiyat > 8 > 13 ... > 233
            vals = list(emas.values())
            is_ideal = (curr > vals[0]) and all(vals[i] > vals[i+1] for i in range(len(vals)-1))
            
            # EMA CROSS: Son gün en az 2 EMA yukarı kesildi mi?
            prev_price = close.iloc[-2]
            prev_emas = {p: calculate_ema(close, p).iloc[-2] for p in ema_periods}
            cross_count = sum(1 for p in ema_periods if prev_price < prev_emas[p] and curr > emas[p])
            
            if not (is_ideal or cross_count >= 2): continue
            
            status = "IDEAL UP" if is_ideal else f"EMA CROSS ({cross_count})"
            
            # 2. KANAL KONTROLÜ (Pearson > 0.80)
            valid_channels = 0
            for per in check_periods:
                y = close.tail(per).values; X = np.arange(len(y)).reshape(-1, 1)
                corr = np.corrcoef(X.flatten(), y)[0, 1]
                if corr > 0.80: valid_channels += 1
                
            if valid_channels < 2: continue
            
            # 3. EXPERTMA PUANLAMA (Entegre)
            score = 0
            inds = [
                close.rolling(173).mean().iloc[-1], # ZLSMA Simüle
                close.rolling(120).mean().iloc[-1], # SMMA
                calculate_ma_type(close, 107).iloc[-1], # MA1
                calculate_ma_type(close, 120).iloc[-1], # MA2
                close.ewm(alpha=0.023).mean().iloc[-1], # M1
                close.rolling(44).quantile(0.89).iloc[-1], # Percentile
                calculate_ema(2*calculate_ema(close, 44)-calculate_ema(close,89), 9).iloc[-1], # FINH Simüle
                calculate_wma(close, 196).iloc[-1], # HMA Simüle
                calculate_ema(close, 50).iloc[-1], # JMA Simüle
                calculate_ma_type(close, 144).iloc[-1], # TEMA
            ]
            for val in inds: 
                if curr > val: score += 1
            
            # Normalize Puan (12 üzerinden)
            final_score = score + 2 # Bonus
            
            if final_score >= 10:
                results.append({
                    'Hisse': hisse, 'Fiyat': curr, 'Statü': status,
                    'Expert Puanı': final_score, 'Kanal Sayısı': valid_channels
                })
        except: continue
        
    if results:
        df_res = pd.DataFrame(results).sort_values(by=['Expert Puanı', 'Kanal Sayısı'], ascending=False)
        out = os.path.join(OUTPUT_DIR, f'Hibrit_V4_FULL_{datetime.now().strftime("%Y%m%d")}.xlsx')
        
        # Renkli Excel
        wb = Workbook(); ws = wb.active; ws.append(list(df_res.columns))
        for r in df_res.values.tolist(): ws.append(r)
        
        header_fill = PatternFill(start_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]: cell.fill = header_fill; cell.font = header_font
        
        wb.save(out)
        print(f"✅ Hibrit Raporu: {out}")

if __name__ == "__main__":
    main()
