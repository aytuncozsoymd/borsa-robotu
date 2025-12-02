import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ==================== BULUT UYUMLU AYARLAR ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'DATAson')
OUTPUT_DIR = BASE_DIR

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'RUA_Trend_Destekli_{datetime.now().strftime("%Y-%m-%d-%H-%M")}.xlsx')

# ==================== YARDIMCI FONKSİYONLAR ====================
def load_stock_df(file_path):
    try:
        df = pd.read_excel(file_path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        col_map = {
            "DATE": ["DATE", "TARIH", "TARİH"], "CLOSING_TL": ["CLOSING_TL", "CLOSE", "KAPANIS"],
            "HIGH_TL": ["HIGH_TL", "HIGH"], "LOW_TL": ["LOW_TL", "LOW"], "VOLUME_TL": ["VOLUME_TL", "VOL"]
        }
        for t, aliases in col_map.items():
            if t not in df.columns:
                for a in aliases:
                    if a in df.columns: df.rename(columns={a: t}, inplace=True); break
        
        if "DATE" in df.columns:
            df["DATE"] = pd.to_datetime(df["DATE"], errors='coerce')
            df.dropna(subset=["DATE", "CLOSING_TL"], inplace=True)
            df.sort_values("DATE", inplace=True)
        
        if "CLOSING_TL" not in df.columns: return None
        return df
    except: return None

def calculate_ema(data, period): return data.ewm(span=period, adjust=False).mean()

def calculate_rsi(series, period=14):
    delta = series.diff()
    gain = (delta.where(delta > 0, 0)).rolling(period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(period).mean()
    rs = gain / loss
    return (100 - (100 / (1 + rs))).fillna(50)

def calculate_mfi(df, period=14):
    if 'VOLUME_TL' not in df.columns or 'HIGH_TL' not in df.columns:
        return calculate_rsi(df['CLOSING_TL'], period) # Volume yoksa RSI dön
    
    tp = (df['HIGH_TL'] + df['LOW_TL'] + df['CLOSING_TL']) / 3
    rmf = tp * df['VOLUME_TL']
    
    pos_flow = np.where(tp > tp.shift(1), rmf, 0)
    neg_flow = np.where(tp < tp.shift(1), rmf, 0)
    
    pos_mf = pd.Series(pos_flow).rolling(period).sum()
    neg_mf = pd.Series(neg_flow).rolling(period).sum()
    
    mfi = 100 - (100 / (1 + (pos_mf / neg_mf)))
    return mfi.fillna(50)

def calculate_bollinger_bands(series, period=20, std_dev=2):
    sma = series.rolling(period).mean()
    std = series.rolling(period).std()
    upper = sma + (std * std_dev)
    lower = sma - (std * std_dev)
    return upper, lower

# ==================== ANA ANALİZ ====================
def main():
    print("RUA v3 + Güçlü Trend Taraması Başlıyor...")
    files = glob.glob(os.path.join(DATA_DIR, '*.xlsx'))
    results = []
    
    for f in files:
        try:
            df = load_stock_df(f)
            if df is None or len(df) < 200: continue
            
            close = df['CLOSING_TL']
            hisse = os.path.basename(f).replace('.xlsx', '')
            
            # --- 1. GÜÇLÜ TREND FİLTRESİ ---
            # Fiyat EMA 200 üzerinde olmalı (Uzun vade trend pozitif)
            ema200 = calculate_ema(close, 200).iloc[-1]
            last_price = close.iloc[-1]
            
            # Pearson Korelasyonu (Son 55 gün - Trend Doğrusallığı)
            y = close.tail(55).values
            x = np.arange(len(y))
            pearson = np.corrcoef(x, y)[0, 1] if len(y) > 1 else 0
            
            # Trend Kriteri: Fiyat > EMA200 VE Pearson > 0 (Pozitif Eğim)
            is_trend_strong = (last_price > ema200) and (pearson > 0)
            
            if not is_trend_strong: continue # Trend yoksa RUA'ya bakma bile
            
            # --- 2. RUA HESAPLAMASI ---
            rsi = calculate_rsi(close)
            mfi = calculate_mfi(df)
            
            rua = (rsi + mfi) / 2
            
            # --- 3. BOLLINGER BANTLARI ---
            bb_upper, bb_lower = calculate_bollinger_bands(rua)
            
            curr_rua = rua.iloc[-1]
            curr_lower = bb_lower.iloc[-1]
            prev_rua = rua.iloc[-2]
            prev_lower = bb_lower.iloc[-2]
            
            # --- 4. SİNYAL MANTIĞI ---
            # AL: RUA, Alt Bandın altında veya Alt bandı yukarı kesiyor
            is_buy = False
            note = ""
            
            # Durum A: RUA Aşırı Satımda (Bandın Altında)
            if curr_rua <= curr_lower:
                is_buy = True
                note = "DİP BÖLGEDE (AL FIRSATI)"
            
            # Durum B: Dönüş Başlamış (Bandı Yukarı Kesti)
            elif prev_rua < prev_lower and curr_rua > curr_lower:
                is_buy = True
                note = "DÖNÜŞ BAŞLADI (TEYİTLİ)"
                
            if is_buy:
                results.append({
                    'Hisse': hisse,
                    'Fiyat': last_price,
                    'Sinyal': note,
                    'RUA Değeri': round(curr_rua, 2),
                    'Trend Durumu': 'GÜÇLÜ (EMA200 Üstü)',
                    'Pearson': round(pearson, 2)
                })
                
        except: continue
        
    if results:
        df_res = pd.DataFrame(results).sort_values(by='RUA Değeri', ascending=True) # En düşük RUA en üstte
        
        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "RUA Trend"
        
        # Başlıklar
        headers = list(df_res.columns)
        ws.append(headers)
        
        # Veriler
        for r in df_res.values.tolist(): ws.append(r)
        
        # Renklendirme
        header_fill = PatternFill(start_color='2C3E50', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        green_fill = PatternFill(start_color='D5F5E3', fill_type='solid') # Dönüş
        yellow_fill = PatternFill(start_color='FCF3CF', fill_type='solid') # Dip
        
        for cell in ws[1]: 
            cell.fill = header_fill; cell.font = white_font
            
        for row in ws.iter_rows(min_row=2):
            val = row[2].value # Sinyal Sütunu
            if "DÖNÜŞ" in val:
                for c in row: c.fill = green_fill
            else:
                for c in row: c.fill = yellow_fill
                
        # Sütun Genişliği
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
            
        wb.save(OUTPUT_FILE)
        print(f"✅ RUA Trend Taraması Tamamlandı: {OUTPUT_FILE}")
        
    else:
        print("Kriterlere uyan hisse bulunamadı (Trendi güçlü olup dip yapan hisse yok).")

if __name__ == "__main__":
    main()