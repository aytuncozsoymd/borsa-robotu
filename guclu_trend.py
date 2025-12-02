import pandas as pd
import numpy as np
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LinearRegression
from scipy.stats import pearsonr
from tqdm import tqdm
import colorama
from colorama import Fore, Style

colorama.init(autoreset=True)

# ==================== BULUT UYUMLU AYARLAR ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'DATAson')
OUTPUT_DIR = BASE_DIR

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)
# ==============================================================

def load_stock_df(file_path):
    try:
        df = pd.read_excel(file_path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        col_map = {"DATE": ["DATE","TARIH"], "CLOSING_TL": ["CLOSING_TL","CLOSE","KAPANIS"], "VOLUME_TL": ["VOLUME_TL","VOL","HACIM"]}
        for t, aliases in col_map.items():
            if t not in df.columns:
                for a in aliases:
                    if a in df.columns: df.rename(columns={a: t}, inplace=True); break
        if "DATE" in df.columns:
            df["DATE"] = pd.to_datetime(df["DATE"], errors='coerce')
            df.dropna(subset=["DATE", "CLOSING_TL"], inplace=True)
            df.sort_values("DATE", inplace=True)
        if "CLOSING_TL" not in df.columns or len(df) < 233: return None
        return df
    except: return None

def calculate_ema(data, period): return data.ewm(span=period, adjust=False).mean()

def calculate_rsi(series, period=14):
    delta = series.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / loss
    return (100 - (100 / (1 + rs))).fillna(50)

def autofit(ws):
    for column in ws.columns:
        length = max(len(str(cell.value)) for cell in column if cell.value)
        ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)

def main():
    print(f"\n🚀 Güçlü Trend (Full Detay) Analizi Başlıyor...")
    files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')]
    
    results_ema = []
    results_strat = []
    
    ema_periods = [8, 13, 21, 34, 55, 89, 144, 233]
    pearson_periods = [55, 89, 144, 233, 377, 610, 987]
    all_pearson = {}

    for file in tqdm(files, desc="Hisseler Taranıyor"):
        df = load_stock_df(os.path.join(DATA_DIR, file))
        if df is None: continue
        
        hisse = file.replace('.xlsx', '')
        close = df['CLOSING_TL']
        curr = close.iloc[-1]
        
        # --- BÖLÜM 1: EMA ANALİZİ ---
        ema_vals = {p: calculate_ema(close, p).iloc[-1] for p in ema_periods}
        
        # IDEAL UP Kontrolü (Sıralı Dizilim)
        vals = list(ema_vals.values())
        is_ideal = all(vals[i] > vals[i+1] for i in range(len(vals)-1)) and (curr > vals[0])
        is_up = all(curr > v for v in vals)
        
        # Pearson Hesaplama
        p_data = {}
        for p in pearson_periods:
            if len(df) >= p:
                y = close.tail(p).values
                X = np.arange(len(y)).reshape(-1, 1)
                p_data[p] = np.corrcoef(X.flatten(), y)[0, 1]
            else: p_data[p] = 0
        
        all_pearson[hisse] = p_data
        
        res_row = {'Hisse': hisse, 'Fiyat': curr, 'Durum': 'IDEAL UP' if is_ideal else ('UP' if is_up else '')}
        for p in ema_periods: res_row[f'EMA{p}'] = round(ema_vals[p], 2)
        for p in pearson_periods: res_row[f'P_{p}'] = round(p_data[p], 2)
        results_ema.append(res_row)
        
        # --- BÖLÜM 2: KANAL VE STRATEJİ ---
        # 233 Günlük Kanal
        if len(df) < 233: continue
        y = close.tail(233).values
        X = np.arange(len(y)).reshape(-1, 1)
        model = LinearRegression().fit(X, y)
        pred = model.predict(X)
        std = np.std(y - pred)
        
        upper = pred[-1] + 2*std
        lower = pred[-1] - 2*std
        dist_down = (curr - lower) / curr * 100
        
        rsi = calculate_rsi(close).iloc[-1]
        vol_surge = False
        if 'VOLUME_TL' in df.columns:
            vol_surge = df['VOLUME_TL'].iloc[-1] > (df['VOLUME_TL'].rolling(10).mean().iloc[-1] * 1.2)
            
        p_233 = p_data[233]
        
        label = ""; score = 0
        if is_ideal and p_233 > 0.90 and rsi < 55 and dist_down < 3:
            label = "🏆 TAM PUANLI"; score = 3
        elif rsi < 35 and vol_surge:
            label = "🚀 TEPKİ ADAYI"; score = 2
        elif is_ideal and p_233 > 0.85 and rsi < 70:
            label = "💪 GÜÇLÜ TREND"; score = 1
            
        if label:
            results_strat.append({
                'Hisse': hisse, 'STRATEJİ': label, 'Fiyat': curr, 'Pearson (233)': round(p_233, 2),
                'RSI': round(rsi, 2), 'Hacim Artışı': "EVET" if vol_surge else "-",
                'Alt Banda Uzaklık %': round(dist_down, 2), 'Skor': score
            })

    # KAYIT
    fname = os.path.join(OUTPUT_DIR, f'Guclu_Trend_FULL_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        if results_strat:
            df_s = pd.DataFrame(results_strat).sort_values(by=['Skor', 'Pearson (233)'], ascending=False).drop(columns=['Skor'])
            df_s.to_excel(writer, sheet_name='Guclu_Trend_Takip', index=False)
            ws = writer.sheets['Guclu_Trend_Takip']
            
            gold = PatternFill(start_color='FFD700', fill_type='solid')
            green = PatternFill(start_color='00B050', fill_type='solid')
            purple = PatternFill(start_color='7030A0', fill_type='solid')
            
            for row in ws.iter_rows(min_row=2):
                val = str(row[1].value)
                if "TAM PUANLI" in val: row[1].fill = gold
                elif "TEPKİ" in val: row[1].fill = green
                elif "GÜÇLÜ" in val: row[1].fill = purple
            autofit(ws)
            
        if results_ema:
            df_e = pd.DataFrame(results_ema)
            df_e.to_excel(writer, sheet_name='EMA_Pearson_Detay', index=False)
            autofit(writer.sheets['EMA_Pearson_Detay'])
            
    print(f"✅ Rapor Kaydedildi: {fname}")

if __name__ == "__main__":
    main()
