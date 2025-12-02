import pandas as pd
import numpy as np
import os
import glob
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LinearRegression
from scipy.stats import pearsonr
from tqdm import tqdm
import colorama
from colorama import Fore, Style

colorama.init(autoreset=True)

# ==================== BULUT UYUMLU AYARLAR ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'DATAson')
OUTPUT_DIR = BASE_DIR

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)
# ==============================================================

# --- YARDIMCI FONKSİYONLAR ---
def load_stock_df(file_path):
    try:
        df = pd.read_excel(file_path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        col_map = {
            "DATE": ["DATE", "TARIH", "TARİH", "TIME"], 
            "CLOSING_TL": ["CLOSING_TL", "CLOSE", "KAPANIS", "KAPANIŞ"],
            "VOLUME_TL": ["VOLUME_TL", "VOL", "HACIM"]
        }
        for t, aliases in col_map.items():
            if t not in df.columns:
                for a in aliases:
                    if a in df.columns: df.rename(columns={a: t}, inplace=True); break
        
        if "DATE" in df.columns:
            df["DATE"] = pd.to_datetime(df["DATE"], errors='coerce')
            df.dropna(subset=["DATE", "CLOSING_TL"], inplace=True)
            df.sort_values("DATE", inplace=True)
            
        if "CLOSING_TL" not in df.columns or len(df) < 55: return None
        return df
    except: return None

def calculate_ema(data, period): return data.ewm(span=period, adjust=False).mean()

def autofit(ws):
    for column in ws.columns:
        length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > length: length = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = min(length + 2, 50)

def add_info(ws, total, filtered):
    r = ws.max_row + 2
    ws.cell(r, 1, "Toplam Taranan:").font = Font(bold=True)
    ws.cell(r, 2, total)
    ws.cell(r+1, 1, "Listedeki Hisse:").font = Font(bold=True)
    ws.cell(r+1, 2, f"{filtered} ({filtered/total:.1%})" if total>0 else 0)

# --- ANA İŞLEM ---
def main():
    print(f"\n🚀 FULL DETAYLI TARAMA (Orijinal Versiyon) Başlıyor...")
    
    # 1. ESKİ RAPORU BUL (Kıyaslama İçin)
    prev_file = None
    all_reports = sorted(glob.glob(os.path.join(OUTPUT_DIR, "Ema_ve_Pearson_Sonuclari-*.xlsx")), key=os.path.getmtime, reverse=True)
    if len(all_reports) > 0:
        prev_file = all_reports[0] # En son oluşturulan dosya (şimdikinden önceki)

    # 2. HİSSELERİ TARA
    files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')]
    
    ema_list = [8, 13, 21, 34, 55, 89, 144, 233]
    pearson_periods = [55, 89, 144, 233, 377, 610, 987]
    
    data_master = [] # Ana veri deposu
    pearson_master = {} # Pearson verileri
    channel_master = [] # Kanal verileri
    
    total_scanned = 0
    
    for f in tqdm(files, desc="Analiz"):
        df = load_stock_df(os.path.join(DATA_DIR, f))
        if df is None: continue
        total_scanned += 1
        hisse = f.replace('.xlsx','')
        close = df['CLOSING_TL']
        curr = close.iloc[-1]
        
        # EMA
        emas = {p: calculate_ema(close, p).iloc[-1] for p in ema_list}
        vals = list(emas.values())
        is_ideal = (curr > vals[0]) and all(vals[i] > vals[i+1] for i in range(len(vals)-1))
        is_up = all(curr > v for v in vals)
        
        status = "IDEAL UP" if is_ideal else ("UP" if is_up else "")
        
        # PEARSON
        p_res = {}
        for p in pearson_periods:
            if len(df) >= p:
                y = close.tail(p).values
                X = np.arange(p).reshape(-1, 1)
                p_res[p] = np.corrcoef(X.flatten(), y)[0, 1]
            else: p_res[p] = np.nan
        pearson_master[hisse] = p_res
        
        # ANA DATA
        row = {'Stock Name': hisse, 'Closing Price': curr, 'Status': status, 'Ideal Status': status} # Orijinal format uyumu
        for p in ema_list: row[f'EMA{p}'] = emas[p]
        data_master.append(row)
        
        # KANAL (233 ve diğerleri için)
        # Sadece liste başı için 233'ü hesaplayıp ekleyelim, detaylı kanal analizini sonra yaparız
        if len(df) >= 233:
            y_ch = close.tail(233).values
            X_ch = np.arange(len(y_ch)).reshape(-1, 1)
            model = LinearRegression().fit(X_ch, y_ch)
            pred = model.predict(X_ch)
            std = np.std(y_ch - pred)
            upper = pred[-1] + 2*std
            lower = pred[-1] - 2*std
            
            diff_up = (upper - curr)/curr*100
            diff_down = (curr - lower)/curr*100
            
            # RSI ve Hacim
            rsi = 50
            try:
                delta = close.diff()
                gain = (delta.where(delta > 0, 0)).rolling(14).mean()
                loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
                rs = gain / loss
                rsi = (100 - (100 / (1 + rs))).fillna(50).iloc[-1]
            except: pass
            
            vol_surge = "HAYIR"
            if "VOLUME_TL" in df.columns:
                v_avg = df['VOLUME_TL'].rolling(10).mean().iloc[-1]
                if df['VOLUME_TL'].iloc[-1] > v_avg * 1.2: vol_surge = "EVET"
            
            channel_master.append({
                'Hisse': hisse, 'Vade': 233, 'Fiyat': curr, 
                'Pearson': p_res[233], 'RSI': rsi, 'Hacim Artışı': vol_surge,
                'Üst Fark %': diff_up, 'Alt Fark %': diff_down
            })

    # --- EXCEL OLUŞTURMA ---
    ts = datetime.now().strftime('%d-%m-%Y-%H-%M')
    fname = os.path.join(OUTPUT_DIR, f'Ema_ve_Pearson_Sonuclari-{ts}.xlsx')
    
    df_main = pd.DataFrame(data_master)
    df_pearson = pd.DataFrame(pearson_master).T
    
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        
        # 1. EMA_Sonuclari
        df_main.to_excel(writer, sheet_name='EMA_Sonuclari', index=False)
        ws_ema = writer.sheets['EMA_Sonuclari']
        # Renklendirme
        blue = PatternFill(start_color='0000FF', fill_type='solid')
        orange = PatternFill(start_color='FFA500', fill_type='solid')
        white = Font(color='FFFFFF', bold=True)
        
        for row in ws_ema.iter_rows(min_row=2):
            st_val = row[10].value # Status
            id_val = row[11].value # Ideal Status
            if st_val == "UP": 
                for c in row: c.fill = blue; c.font = white
            if id_val == "IDEAL UP":
                row[11].fill = orange; row[11].font = white
        autofit(ws_ema)
        
        # 2. UP-Ideal UP
        ws_up = writer.book.create_sheet("UP-Ideal UP")
        ws_up.append(["Hisse", "Durum", "Fiyat"])
        for r in data_master:
            if r['Status'] in ["UP", "IDEAL UP"]:
                ws_up.append([r['Stock Name'], r['Status'], r['Closing Price']])
        autofit(ws_up)
        
        # 3. Pearson_Sonuclari
        df_pearson.to_excel(writer, sheet_name='Pearson_Sonuclari')
        autofit(writer.sheets['Pearson_Sonuclari'])
        
        # 4. En_Yuksek_Pearson
        # (Basitleştirilmiş)
        ws_high = writer.book.create_sheet("En_Yuksek_Pearson")
        ws_high.append(["Hisse", "En Yüksek Periyot", "Değer"])
        for h, vals in pearson_master.items():
            best_p = max(vals, key=vals.get)
            ws_high.append([h, best_p, vals[best_p]])
        autofit(ws_high)
        
        # 5. Pozitif_Pearson
        pos_list = {k: v for k, v in pearson_master.items() if all(val > 0 for val in v.values() if pd.notna(val))}
        pd.DataFrame(pos_list).T.to_excel(writer, sheet_name='Pozitif_Pearson')
        autofit(writer.sheets['Pozitif_Pearson'])
        
        # 6. Com144-233-377 (Kesişim)
        com_list = []
        for h, vals in pearson_master.items():
            if vals.get(144,0)>0.8 and vals.get(233,0)>0.8 and vals.get(377,0)>0.8:
                com_list.append({'Hisse': h, '144': vals[144], '233': vals[233], '377': vals[377]})
        pd.DataFrame(com_list).to_excel(writer, sheet_name='Com144-233-377', index=False)
        
        # 7-11. Period Ideal Up (Her periyot için ayrı sayfa)
        target_periods = [144, 233, 377, 610, 987]
        for p in target_periods:
            p_ideal_list = []
            for r in data_master:
                h = r['Stock Name']
                if r['Ideal Status'] == "IDEAL UP" and pearson_master[h].get(p, 0) > 0.85:
                    p_ideal_list.append({'Hisse': h, 'Fiyat': r['Closing Price'], f'Pearson_{p}': pearson_master[h][p]})
            
            if p_ideal_list:
                pd.DataFrame(p_ideal_list).to_excel(writer, sheet_name=f'{p}IdealUp', index=False)
                autofit(writer.sheets[f'{p}IdealUp'])

        # 12. Rapor_Upd (Karşılaştırma)
        ws_rep = writer.book.create_sheet("Rapor_Upd")
        ws_rep.append(["Durum", "Hisse", "Fiyat"])
        
        current_ideal = {r['Stock Name'] for r in data_master if r['Ideal Status'] == "IDEAL UP"}
        prev_ideal = set()
        
        if prev_file:
            try:
                df_prev = pd.read_excel(prev_file, sheet_name='EMA_Sonuclari')
                prev_ideal = set(df_prev[df_prev['Ideal Status'] == 'IDEAL UP']['Stock Name'])
            except: pass
            
        new_entries = current_ideal - prev_ideal
        dropped = prev_ideal - current_ideal
        
        for h in new_entries: ws_rep.append(["YENİ GİREN", h, ""])
        for h in dropped: ws_rep.append(["ÇIKAN", h, ""])
        if not prev_file: ws_rep.append(["Bilgi", "Geçmiş dosya bulunamadı, hepsi yeni kabul edildi.", ""])
        autofit(ws_rep)
        
        # 13. Kanal_Ekstra
        df_ch = pd.DataFrame(channel_master)
        if not df_ch.empty:
            df_ch.to_excel(writer, sheet_name='Kanal_Ekstra', index=False)
            autofit(writer.sheets['Kanal_Ekstra'])
            
            # 14. ListeBaşı (Filtreli)
            # Kriter: Pearson > 0.90 VE (Alt bant %2 yakın VEYA Üst bant %2 yakın)
            lb = df_ch[(df_ch['Pearson'] > 0.90) & ((df_ch['Alt Fark %'] <= 2) | (df_ch['Üst Fark %'] <= 2))]
            if not lb.empty:
                lb.to_excel(writer, sheet_name='ListeBaşı', index=False)
                ws_lb = writer.sheets['ListeBaşı']
                red = PatternFill(start_color='FF0000', fill_type='solid')
                navy = PatternFill(start_color='000080', fill_type='solid')
                white = Font(color='FFFFFF', bold=True)
                
                for row in ws_lb.iter_rows(min_row=2):
                    for cell in row: cell.fill = navy; cell.font = white
                    # Alt banda yakınsa Kırmızı
                    if row[7].value <= 2: # Alt Fark %
                        for cell in row: cell.fill = red
                autofit(ws_lb)

    print(f"✅ Full Rapor Hazırlandı: {fname}")

if __name__ == "__main__":
    main()
