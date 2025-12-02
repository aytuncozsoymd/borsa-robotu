import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'DATAson')
OUTPUT_DIR = BASE_DIR

if not os.path.exists(DATA_DIR): os.makedirs(DATA_DIR)

def load_stock_df(path):
    try:
        df = pd.read_excel(path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        if "CLOSING_TL" in df.columns: return df
        if "CLOSE" in df.columns: return df.rename(columns={"CLOSE": "CLOSING_TL", "VOLUME": "VOLUME_TL"})
        return None
    except: return None

def main():
    print("Hacimli EMA Cross (Full) Taraması...")
    files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')]
    results = []
    
    ema_periods = [8, 13, 21, 34, 55, 89, 144, 233]
    
    for f in files:
        df = load_stock_df(os.path.join(DATA_DIR, f))
        if df is None or len(df) < 240: continue
        
        close = df['CLOSING_TL']
        vol = df.get('VOLUME_TL', pd.Series(0, index=df.index))
        
        curr = close.iloc[-1]; prev = close.iloc[-2]
        
        # Kırılımları Say
        crosses = []
        for p in ema_periods:
            ema = close.ewm(span=p, adjust=False).mean()
            if prev < ema.iloc[-2] and curr > ema.iloc[-1]:
                crosses.append(str(p))
                
        if crosses:
            # Hacim Kontrolü
            vol_avg = vol.rolling(20).mean().iloc[-1]
            vol_curr = vol.iloc[-1]
            is_vol = vol_curr > (vol_avg * 1.2)
            pct_change = ((vol_curr - vol_avg)/vol_avg)*100 if vol_avg > 0 else 0
            
            results.append({
                'Hisse': f.replace('.xlsx',''),
                'Fiyat': curr,
                'Kırılan EMA Sayısı': len(crosses),
                'Hacim Durumu': "HACİMLİ" if is_vol else "-",
                'Hacim Değişimi %': round(pct_change, 1),
                'Kırılanlar': ",".join(crosses)
            })
            
    if results:
        df_res = pd.DataFrame(results).sort_values(by=['Kırılan EMA Sayısı', 'Hacim Değişimi %'], ascending=False)
        out = os.path.join(OUTPUT_DIR, f'EMA_Cross_Full_{datetime.now().strftime("%Y%m%d")}.xlsx')
        
        wb = Workbook(); ws = wb.active
        ws.append(list(df_res.columns))
        for r in df_res.values.tolist(): ws.append(r)
        
        # Renklendirme
        green_fill = PatternFill(start_color='D5F5E3', fill_type='solid')
        header_fill = PatternFill(start_color='2C3E50', fill_type='solid')
        font_white = Font(color='FFFFFF', bold=True)
        font_green = Font(color='006400', bold=True)
        
        for cell in ws[1]: cell.fill = header_fill; cell.font = font_white
        
        for row in ws.iter_rows(min_row=2):
            if row[3].value == "HACİMLİ": # Hacim sütunu
                row[3].fill = green_fill; row[3].font = font_green
                
        wb.save(out)
        print(f"✅ Dosya Kaydedildi: {out}")

if __name__ == "__main__":
    main()
